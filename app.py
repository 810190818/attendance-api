"""
鑰冨嫟璇嗗埆 App - Flask 鍚庣
"""

import os
import json
import sqlite3
import base64
import uuid
import logging
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS

app = Flask(__name__)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB 涓婁紶闄愬埗

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'attendance.db')
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
LOG_DIR = os.path.join(BASE_DIR, 'logs')

for d in [UPLOAD_DIR, LOG_DIR]:
    os.makedirs(d, exist_ok=True)

# 鏃ュ織閰嶇疆
LOG_FILE = os.path.join(LOG_DIR, f'app_{datetime.now().strftime("%Y%m%d")}.log')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€
# 閿欒浠ｇ爜瀹氫箟
# 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€
ERROR_CODES = {
    'E001': '鍥剧墖涓婁紶澶辫触',
    'E002': '鏂囦欢鏍煎紡涓嶆敮鎸?,
    'E003': 'MiniMax API 璇锋眰澶辫触',
    'E004': 'MiniMax API 杩斿洖鏍煎紡閿欒',
    'E005': '鍥剧墖缂栫爜澶辫触',
    'E006': '鏁版嵁搴撳啓鍏ュけ璐?,
    'E007': '璁板綍涓嶅瓨鍦?,
    'E008': '鍙傛暟缂哄け',
    'E009': 'Excel 瀵煎嚭澶辫触',
    'E010': '鍘嗗彶璁板綍涓嶅瓨鍦?,
}


def log(code, message, extra=None):
    log_data = {'code': code, 'message': message}
    if extra:
        log_data.update(extra)
    logger.info(json.dumps(log_data, ensure_ascii=False))


def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS attendance_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            record_id TEXT UNIQUE NOT NULL,
            created_at TEXT NOT NULL,
            image_path TEXT,
            status TEXT DEFAULT 'pending',
            raw_result TEXT,
            confirmed_data TEXT,
            employee_count INTEGER DEFAULT 0,
            remark TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            employee_id TEXT,
            department TEXT,
            created_at TEXT NOT NULL
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS operation_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            record_id TEXT,
            operation TEXT NOT NULL,
            operator TEXT DEFAULT 'system',
            detail TEXT,
            created_at TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()
    log('INFO', '鏁版嵁搴撳垵濮嬪寲瀹屾垚', {'db_path': DB_PATH})


# 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€
# MiniMax 瑙嗚璇嗗埆
# 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€
MINIMAX_API_URL = 'https://api.minimax.chat/v1/text/chatcompletion_v2'
MINIMAX_API_KEY = os.environ.get('MINIMAX_API_KEY', '')
MINIMAX_MODEL = os.environ.get('MINIMAX_MODEL', 'MiniMax-M2.7-highspeed')

ATTENDANCE_PROMPT = """浣犳槸涓€涓€冨嫟鏁版嵁璇嗗埆涓撳銆傝浠旂粏璇嗗埆鍥剧墖涓殑鑰冨嫟璁板綍銆?
銆愯瘑鍒姹傘€?1. 蹇呴』閫愭潯璇嗗埆锛岀姝㈠悎骞朵换浣曡褰?2. 姣忔潯璁板綍鍖呭惈锛氬鍚嶃€佸伐鍙枫€佹棩鏈燂紙YYYY-MM-DD鏍煎紡锛夈€佷笂鐝椂闂达紙HH:MM鏍煎紡锛夈€佷笅鐝椂闂达紙HH:MM鏍煎紡锛夈€佽€冨嫟鐘舵€?3. 鏃ユ湡蹇呴』鏄?YYYY-MM-DD 鏍煎紡
4. 鏃堕棿蹇呴』鏄?HH:MM 鏍煎紡锛堝 09:00銆?8:30锛?5. 鑰冨嫟鐘舵€佹灇涓撅細姝ｅ父銆佽繜鍒般€佹棭閫€銆佺己鍕ゃ€佸姞鐝€佽皟浼戙€佺梾鍋囥€佷簨鍋囥€佸鍋囥€佷骇鍋?6. 濡傛灉鏌愬瓧娈垫棤娉曡瘑鍒紝鏍囨敞涓?"鏈煡"
7. 鍥剧墖涓鏈夊涓汉鍛樿褰曪紝璇峰叏閮ㄨ瘑鍒嚭鏉?8. 杩斿洖 JSON 鏁扮粍鏍煎紡锛屾瘡鏉¤褰曚竴涓璞?
銆愯繑鍥炴牸寮忋€?[
  {
    "seq": 1,
    "name": "寮犱笁",
    "employee_id": "1001",
    "date": "2024-03-01",
    "weekday": "鏄熸湡浜?,
    "check_in": "09:00",
    "check_out": "18:00",
    "status": "姝ｅ父",
    "confidence": 0.95,
    "remark": ""
  },
  ...
]

璇风洿鎺ヨ繑鍥?JSON锛屼笉瑕佹湁浠讳綍鍏朵粬鏂囧瓧銆?""


def call_minimax_vision(image_base64: str) -> dict:
    import urllib.request
    import urllib.error

    if not MINIMAX_API_KEY:
        raise Exception('MiniMax API Key 鏈厤缃?)

    payload = {
        'model': MINIMAX_MODEL,
        'messages': [
            {
                'role': 'user',
                'content': [
                    {
                        'type': 'image_url',
                        'image_url': {
                            'url': f'data:image/jpeg;base64,{image_base64}'
                        }
                    },
                    {
                        'type': 'text',
                        'text': ATTENDANCE_PROMPT
                    }
                ]
            }
        ],
        'temperature': 0.1
    }

    data = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(
        MINIMAX_API_URL,
        data=data,
        headers={
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {MINIMAX_API_KEY}'
        },
        method='POST'
    )

    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read().decode('utf-8'))
            log('INFO', 'MiniMax API 璋冪敤鎴愬姛', {'model': MINIMAX_MODEL})
            return result
    except urllib.error.HTTPError as e:
        error_body = e.read().decode('utf-8', errors='replace')
        log('E003', f'MiniMax API HTTP閿欒', {'status_code': e.code, 'error': error_body[:500]})
        raise Exception(f'MiniMax API 璇锋眰澶辫触: HTTP {e.code}')
    except Exception as e:
        log('E003', f'MiniMax API 璋冪敤寮傚父', {'error': str(e)})
        raise


def parse_minimax_response(response: dict) -> list:
    try:
        choices = response.get('choices', [])
        if not choices:
            raise Exception('MiniMax 杩斿洖鏍煎紡閿欒锛氭棤 choices')

        content = choices[0].get('message', {}).get('content', '')
        content = content.strip()
        if content.startswith('```'):
            lines = content.split('\n')
            content = '\n'.join(lines[1:-1]) if lines[-1] == '```' else '\n'.join(lines[1:])
        elif content.startswith('['):
            pass
        else:
            start = content.find('[')
            end = content.rfind(']')
            if start != -1 and end != -1:
                content = content[start:end+1]

        records = json.loads(content)
        if not isinstance(records, list):
            raise Exception('杩斿洖涓嶆槸鏁扮粍鏍煎紡')

        log('INFO', f'MiniMax 瑙ｆ瀽鎴愬姛锛屽叡 {len(records)} 鏉¤褰?)
        return records
    except json.JSONDecodeError as e:
        log('E004', f'MiniMax JSON 瑙ｆ瀽澶辫触', {'error': str(e)})
        raise Exception(f'MiniMax 杩斿洖鏍煎紡閿欒锛欽SON瑙ｆ瀽澶辫触')


# 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€
# API 璺敱
# 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')


@app.route('/api/upload', methods=['POST'])
def upload_image():
    try:
        log('INFO', '鏀跺埌鍥剧墖涓婁紶璇锋眰')

        if 'image' not in request.files and 'image' not in request.form:
            return jsonify({'code': 'E008', 'msg': '璇锋彁渚涘浘鐗?, 'data': None}), 400

        if 'image' in request.files:
            file = request.files['image']
            image_data = file.read()
        else:
            image_data = base64.b64decode(request.form['image'])

        if len(image_data) < 1000:
            return jsonify({'code': 'E002', 'msg': '鏂囦欢鏍煎紡涓嶆敮鎸佹垨鏂囦欢鎹熷潖', 'data': None}), 400

        record_id = str(uuid.uuid4())[:8]
        filename = f'{record_id}.jpg'
        filepath = os.path.join(UPLOAD_DIR, filename)
        with open(filepath, 'wb') as f:
            f.write(image_data)
        log('INFO', f'鍥剧墖宸蹭繚瀛?, {'record_id': record_id})

        image_b64 = base64.b64encode(image_data).decode('utf-8')

        try:
            response = call_minimax_vision(image_b64)
            records = parse_minimax_response(response)
        except Exception as e:
            return jsonify({'code': 'E003', 'msg': f'璇嗗埆澶辫触锛歿str(e)}', 'data': None}), 500

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        try:
            c.execute('''
                INSERT INTO attendance_records (record_id, created_at, image_path, status, raw_result, employee_count)
                VALUES (?, ?, ?, 'recognized', ?, ?)
            ''', (record_id, datetime.now().isoformat(), filepath, json.dumps(records, ensure_ascii=False), len(records)))
            c.execute('''
                INSERT INTO operation_logs (record_id, operation, detail, created_at)
                VALUES (?, 'upload', ?, ?)
            ''', (record_id, json.dumps({'action': 'upload', 'employee_count': len(records)}, ensure_ascii=False), datetime.now().isoformat()))
            conn.commit()
        except Exception as e:
            log('E006', '鏁版嵁搴撳啓鍏ュけ璐?, {'error': str(e)})
            conn.rollback()
            return jsonify({'code': 'E006', 'msg': '鏁版嵁搴撳啓鍏ュけ璐?, 'data': None}), 500
        finally:
            conn.close()

        return jsonify({
            'code': '0000',
            'msg': '璇嗗埆鎴愬姛',
            'data': {'record_id': record_id, 'records': records, 'count': len(records)}
        })

    except Exception as e:
        log('E001', f'鍥剧墖涓婁紶寮傚父', {'error': str(e)})
        return jsonify({'code': 'E001', 'msg': f'鍥剧墖涓婁紶澶辫触锛歿str(e)}', 'data': None}), 500


@app.route('/api/records/<record_id>', methods=['GET'])
def get_record(record_id):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('SELECT * FROM attendance_records WHERE record_id = ?', (record_id,))
    row = c.fetchone()
    conn.close()

    if not row:
        return jsonify({'code': 'E007', 'msg': '璁板綍涓嶅瓨鍦?, 'data': None}), 404

    raw_result = json.loads(row['raw_result']) if row['raw_result'] else []
    confirmed_data = json.loads(row['confirmed_data']) if row['confirmed_data'] else raw_result

    return jsonify({
        'code': '0000',
        'msg': '鎴愬姛',
        'data': {
            'record_id': row['record_id'],
            'created_at': row['created_at'],
            'status': row['status'],
            'records': confirmed_data,
            'employee_count': row['employee_count'],
            'remark': row['remark']
        }
    })


@app.route('/api/records/<record_id>', methods=['PUT'])
def update_record(record_id):
    data = request.get_json()
    records = data.get('records', [])
    remark = data.get('remark', '')

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute('''
            UPDATE attendance_records
            SET confirmed_data = ?, status = 'confirmed', remark = ?
            WHERE record_id = ?
        ''', (json.dumps(records, ensure_ascii=False), remark, record_id))
        c.execute('''
            INSERT INTO operation_logs (record_id, operation, detail, created_at)
            VALUES (?, 'update', ?, ?)
        ''', (record_id, json.dumps({'action': 'confirm', 'count': len(records)}, ensure_ascii=False), datetime.now().isoformat()))
        conn.commit()
    except Exception as e:
        return jsonify({'code': 'E006', 'msg': '鏇存柊澶辫触', 'data': None}), 500
    finally:
        conn.close()

    return jsonify({'code': '0000', 'msg': '鏇存柊鎴愬姛', 'data': {'record_id': record_id}})


@app.route('/api/records/<record_id>', methods=['DELETE'])
def delete_record(record_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT image_path FROM attendance_records WHERE record_id = ?', (record_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'code': 'E010', 'msg': '璁板綍涓嶅瓨鍦?, 'data': None}), 404

    c.execute('DELETE FROM attendance_records WHERE record_id = ?', (record_id,))
    c.execute('DELETE FROM operation_logs WHERE record_id = ?', (record_id,))
    conn.commit()
    conn.close()

    if row[0] and os.path.exists(row[0]):
        os.remove(row[0])

    return jsonify({'code': '0000', 'msg': '鍒犻櫎鎴愬姛', 'data': None})


@app.route('/api/records', methods=['GET'])
def list_records():
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 20))
    offset = (page - 1) * page_size

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM attendance_records')
    total = c.fetchone()[0]
    c.execute('''
        SELECT record_id, created_at, status, employee_count, remark
        FROM attendance_records
        ORDER BY created_at DESC
        LIMIT ? OFFSET ?
    ''', (page_size, offset))
    rows = c.fetchall()
    conn.close()

    return jsonify({
        'code': '0000',
        'msg': '鎴愬姛',
        'data': {'total': total, 'page': page, 'page_size': page_size, 'records': [dict(r) for r in rows]}
    })


@app.route('/api/export/<record_id>', methods=['GET'])
def export_excel(record_id):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'code': 'E009', 'msg': 'Excel 搴撴湭瀹夎', 'data': None}), 500

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('SELECT raw_result, confirmed_data FROM attendance_records WHERE record_id = ?', (record_id,))
    row = c.fetchone()
    conn.close()

    if not row:
        return jsonify({'code': 'E010', 'msg': '璁板綍涓嶅瓨鍦?, 'data': None}), 404

    data = json.loads(row['confirmed_data']) if row['confirmed_data'] else json.loads(row['raw_result'])
    if not data:
        data = []

    wb = Workbook()
    ws = wb.active
    ws.title = '鑰冨嫟璁板綍'

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='寰蒋闆呴粦', bold=True, color='FFFFFF', size=12)
    cell_font = Font(name='寰蒋闆呴粦', size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    headers = ['搴忓彿', '濮撳悕', '宸ュ彿', '鏃ユ湡', '鏄熸湡', '涓婄彮鏃堕棿', '涓嬬彮鏃堕棿', '鑰冨嫟鐘舵€?, '澶囨敞']
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, record in enumerate(data, 1):
        confidence = record.get('confidence', 1.0)
        ws.append([
            i,
            record.get('name', ''),
            record.get('employee_id', ''),
            record.get('date', ''),
            record.get('weekday', ''),
            record.get('check_in', ''),
            record.get('check_out', ''),
            record.get('status', ''),
            record.get('remark', '')
        ])
        for col_idx in range(1, 10):
            cell = ws.cell(row=i+1, column=col_idx)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border

    col_widths = [8, 12, 12, 14, 10, 12, 12, 12, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    export_filename = f'attendance_{record_id}_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    export_path = os.path.join(BASE_DIR, 'exports', export_filename)
    os.makedirs(os.path.dirname(export_path), exist_ok=True)
    wb.save(export_path)

    log('INFO', 'Excel 瀵煎嚭鎴愬姛', {'record_id': record_id})
    return jsonify({'code': '0000', 'msg': '瀵煎嚭鎴愬姛', 'data': {'download_url': f'/api/download/{export_filename}'}})


@app.route('/api/download/<filename>')
def download_file(filename):
    return send_from_directory(os.path.join(BASE_DIR, 'exports'), filename)


@app.route('/api/employees', methods=['GET'])
def list_employees():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('SELECT * FROM employees ORDER BY name')
    rows = c.fetchall()
    conn.close()
    return jsonify({'code': '0000', 'msg': '鎴愬姛', 'data': [dict(r) for r in rows]})


@app.route('/api/employees', methods=['POST'])
def add_employee():
    data = request.get_json()
    name = data.get('name', '').strip()
    employee_id = data.get('employee_id', '').strip()
    department = data.get('department', '').strip()

    if not name:
        return jsonify({'code': 'E008', 'msg': '濮撳悕涓嶈兘涓虹┖', 'data': None}), 400

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute('''
            INSERT OR REPLACE INTO employees (name, employee_id, department, created_at)
            VALUES (?, ?, ?, ?)
        ''', (name, employee_id, department, datetime.now().isoformat()))
        conn.commit()
    except Exception as e:
        return jsonify({'code': 'E006', 'msg': str(e), 'data': None}), 500
    finally:
        conn.close()

    return jsonify({'code': '0000', 'msg': '娣诲姞鎴愬姛', 'data': {'name': name}})


@app.route('/api/employees/match', methods=['POST'])
def match_employee():
    data = request.get_json()
    names = data.get('names', [])

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    result = {}
    for name in names:
        c.execute('SELECT employee_id FROM employees WHERE name = ?', (name,))
        row = c.fetchone()
        result[name] = row['employee_id'] if row else None
    conn.close()

    return jsonify({'code': '0000', 'msg': '鎴愬姛', 'data': result})


@app.route('/api/logs', methods=['GET'])
def get_logs():
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 50))
    offset = (page - 1) * page_size
    record_id = request.args.get('record_id', None)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    if record_id:
        c.execute('SELECT COUNT(*) FROM operation_logs WHERE record_id = ?', (record_id,))
        total = c.fetchone()[0]
        c.execute('SELECT * FROM operation_logs WHERE record_id = ? ORDER BY created_at DESC LIMIT ? OFFSET ?', (record_id, page_size, offset))
    else:
        c.execute('SELECT COUNT(*) FROM operation_logs')
        total = c.fetchone()[0]
        c.execute('SELECT * FROM operation_logs ORDER BY created_at DESC LIMIT ? OFFSET ?', (page_size, offset))

    rows = c.fetchall()
    conn.close()

    return jsonify({'code': '0000', 'msg': '鎴愬姛', 'data': {'total': total, 'page': page, 'page_size': page_size, 'logs': [dict(r) for r in rows]}})


@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'code': '0000', 'msg': 'OK', 'data': {'status': 'running'}})


# 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€
# 鍚姩
# 鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€鈹€
init_db()

if __name__ == '__main__':
    log('INFO', '鑰冨嫟璇嗗埆鏈嶅姟鍚姩', {'port': 5124})
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5124)), debug=False)
